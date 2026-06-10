"""
Panther v2 - AutoCAD Layout Generator
Reads coordinates from Excel (PASTE sheet), determines orientation,
assigns correct callout viewport, generates AutoLISP, optionally injects into AutoCAD.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import tempfile
from dataclasses import dataclass, field
from typing import Optional
from pathlib import Path


# ──────────────────────────────────────────────
# Data model
# ──────────────────────────────────────────────

@dataclass
class LayoutRow:
    row_num: int
    # Main viewport (mandatory)
    mx1: float = 0.0
    my1: float = 0.0
    mz1: float = 0.0
    mx2: float = 0.0
    my2: float = 0.0
    mz2: float = 0.0
    # Horizontal callout
    hx1: float = 0.0
    hy1: float = 0.0
    hz1: float = 0.0
    hx2: float = 0.0
    hy2: float = 0.0
    hz2: float = 0.0
    # Vertical callout
    vx1: float = 0.0
    vy1: float = 0.0
    vz1: float = 0.0
    vx2: float = 0.0
    vy2: float = 0.0
    vz2: float = 0.0
    # Derived
    has_horizontal: bool = False
    has_vertical: bool = False

    @property
    def is_horizontal(self) -> bool:
        """Main rectangle is wider than it is tall."""
        return abs(self.mx2 - self.mx1) >= abs(self.my2 - self.my1)

    @property
    def callout_coords(self) -> Optional[tuple]:
        """Return (x1,y1,x2,y2) for the callout viewport, or None."""
        if self.is_horizontal and self.has_horizontal:
            return (self.hx1, self.hy1, self.hx2, self.hy2)
        elif not self.is_horizontal and self.has_vertical:
            return (self.vx1, self.vy1, self.vx2, self.vy2)
        return None

    def paper_viewport(self, x1, y1, x2, y2, scale: float,
                       offset_x: float = 0, offset_y: float = 0):
        """
        Convert model-space extents to paper-space viewport rectangle.
        Paper origin offset lets you place multiple viewports side-by-side.
        Returns (px1, py1, px2, py2) in mm.
        """
        w = abs(x2 - x1) / scale
        h = abs(y2 - y1) / scale
        px1 = offset_x
        py1 = offset_y
        px2 = offset_x + w
        py2 = offset_y + h
        return (px1, py1, px2, py2)


# ──────────────────────────────────────────────
# LISP Generator
# ──────────────────────────────────────────────

class LispGenerator:

    @staticmethod
    def _fmt(v: float) -> str:
        return f"{v:.6f}"

    @staticmethod
    def _pt(x: float, y: float) -> str:
        return f"{LispGenerator._fmt(x)},{LispGenerator._fmt(y)}"

    def generate(self, rows: list[LayoutRow], scale: float,
                 prefix: str = "Layout") -> str:
        lines = [
            "; Panther v2 - AutoCAD Layout Generator",
            f"; Scale 1:{int(scale) if scale == int(scale) else scale}",
            "; Auto-generated - do not edit manually",
            "",
            "(defun c:panther (/ doc)",
            "  (setvar \"CMDECHO\" 0)",
            "  (setq doc (vla-get-ActiveDocument (vlax-get-acad-object)))",
            "",
        ]

        for row in rows:
            name = f"{prefix}_{row.row_num}"
            # Paper viewport for main
            pw = abs(row.mx2 - row.mx1) / scale
            ph = abs(row.my2 - row.my1) / scale

            # Place main viewport starting at paper 0,0
            mpx1, mpy1 = 0.0, 0.0
            mpx2, mpy2 = pw, ph

            # Callout offset: place to the right of main with 10mm gap
            callout = row.callout_coords
            if callout:
                cx1, cy1, cx2, cy2 = callout
                cw = abs(cx2 - cx1) / scale
                ch = abs(cy2 - cy1) / scale
                cpx1 = pw + 10.0
                cpy1 = 0.0
                cpx2 = cpx1 + cw
                cpy2 = ch

            lines.append(f"  ; --- Layout {row.row_num}: {name} ---")
            lines.append(f"  (command \"_LAYOUT\" \"COPY\" \"Model\" \"{name}\")")
            lines.append(f"  (command \"_CTAB\" \"{name}\")")
            lines.append("")

            # Main viewport
            lines.append(f"  ; Main viewport (model: {LispGenerator._pt(row.mx1,row.my1)} -> {LispGenerator._pt(row.mx2,row.my2)})")
            lines.append(f"  (command \"_MVIEW\" \"{LispGenerator._pt(mpx1,mpy1)}\" \"{LispGenerator._pt(mpx2,mpy2)}\")")
            # Set viewport scale
            lines.append(f"  (command \"_ZOOM\" \"E\")")
            lines.append(f"  (command \"_ZOOM\" \"{LispGenerator._fmt(1.0/scale)}XP\")")
            lines.append(f"  ; Set view center to model extents center")
            cx_m = (row.mx1 + row.mx2) / 2
            cy_m = (row.my1 + row.my2) / 2
            lines.append(f"  (command \"_PAN\" \"{LispGenerator._pt(cx_m, cy_m)}\" \"\")")
            lines.append("")

            # Callout viewport
            if callout:
                cx1, cy1, cx2, cy2 = callout
                cx_c = (cx1 + cx2) / 2
                cy_c = (cy1 + cy2) / 2
                orient_label = "H" if row.is_horizontal else "V"
                lines.append(f"  ; Callout viewport [{orient_label}] (model: {LispGenerator._pt(cx1,cy1)} -> {LispGenerator._pt(cx2,cy2)})")
                lines.append(f"  (command \"_MVIEW\" \"{LispGenerator._pt(cpx1,cpy1)}\" \"{LispGenerator._pt(cpx2,cpy2)}\")")
                lines.append(f"  (command \"_ZOOM\" \"E\")")
                lines.append(f"  (command \"_ZOOM\" \"{LispGenerator._fmt(1.0/scale)}XP\")")
                lines.append(f"  (command \"_PAN\" \"{LispGenerator._pt(cx_c, cy_c)}\" \"\")")
                lines.append("")

            # Lock viewports and return to paper space
            lines.append(f"  (command \"_MSPACE\")")
            lines.append(f"  (command \"_VPLAYER\" \"LOCK\" \"ON\" \"ALL\" \"\")")
            lines.append(f"  (command \"_PSPACE\")")
            lines.append("")

        lines.extend([
            "  (setvar \"CMDECHO\" 1)",
            "  (princ (strcat \"\\nPanther: \" (itoa " + str(len(rows)) + ") \" layout(s) created.\"))",
            "  (princ)",
            ")",
            "",
            "; Auto-run",
            "(c:panther)",
        ])

        return "\n".join(lines)


# ──────────────────────────────────────────────
# Excel Reader
# ──────────────────────────────────────────────

class ExcelReader:
    """
    Reads the PASTE sheet where each cell holds a comma-separated 'X,Y,Z' string.

    Column layout (0-indexed):
      A (0) = main  X1,Y1,Z1   B (1) = main  X2,Y2,Z2   C (2) = blank
      D (3) = horiz X1,Y1,Z1   E (4) = horiz X2,Y2,Z2   F (5) = blank
      G (6) = vert  X1,Y1,Z1   H (7) = vert  X2,Y2,Z2
    """

    @staticmethod
    def _parse_xyz(cell_val) -> Optional[tuple]:
        """Parse 'X,Y,Z' or 'X,Y' from a single cell. Returns (x, y, z) or None."""
        if cell_val is None:
            return None
        s = str(cell_val).strip()
        if not s:
            return None
        parts = [p.strip() for p in s.split(',')]
        try:
            x = float(parts[0])
            y = float(parts[1]) if len(parts) > 1 else None
            z = float(parts[2]) if len(parts) > 2 else 0.0
            if y is None:
                return None
            return (x, y, z)
        except (ValueError, IndexError):
            return None

    @staticmethod
    def read_csv(filepath: str) -> list[LayoutRow]:
        """
        Read a CSV where each row is: A, B, (blank), D, E, (blank), G, H
        matching the same column layout as the PASTE sheet.
        Each cell value is 'X,Y,Z' or 'X,Y'.
        """
        import csv as csv_mod
        rows_out = []
        parse = ExcelReader._parse_xyz
        with open(filepath, newline='', encoding='utf-8-sig') as f:
            reader = csv_mod.reader(f)
            headers_skipped = False
            row_num = 0
            for raw in reader:
                # Skip first row if it looks like a header (non-numeric first cell)
                if not headers_skipped:
                    headers_skipped = True
                    first = raw[0].strip() if raw else ""
                    try:
                        float(first.split(',')[0])
                    except ValueError:
                        continue  # it's a header row

                row_num += 1
                if not any(v.strip() for v in raw):
                    continue

                main1 = parse(raw[0] if len(raw) > 0 else None)
                main2 = parse(raw[1] if len(raw) > 1 else None)
                if main1 is None or main2 is None:
                    continue

                lr = LayoutRow(row_num=row_num,
                               mx1=main1[0], my1=main1[1], mz1=main1[2],
                               mx2=main2[0], my2=main2[1], mz2=main2[2])

                h1 = parse(raw[3] if len(raw) > 3 else None)
                h2 = parse(raw[4] if len(raw) > 4 else None)
                if h1 and h2:
                    lr.hx1, lr.hy1, lr.hz1 = h1
                    lr.hx2, lr.hy2, lr.hz2 = h2
                    lr.has_horizontal = True

                v1 = parse(raw[6] if len(raw) > 6 else None)
                v2 = parse(raw[7] if len(raw) > 7 else None)
                if v1 and v2:
                    lr.vx1, lr.vy1, lr.vz1 = v1
                    lr.vx2, lr.vy2, lr.vz2 = v2
                    lr.has_vertical = True

                rows_out.append(lr)
        return rows_out

    @staticmethod
    def read_paste_sheet(filepath: str) -> list[LayoutRow]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required.  Run: pip install openpyxl")

        wb = openpyxl.load_workbook(filepath, data_only=True)

        sheet_name = None
        for name in wb.sheetnames:
            if name.strip().upper() == "PASTE":
                sheet_name = name
                break
        if not sheet_name:
            raise ValueError(f"No 'PASTE' sheet found. Available sheets: {wb.sheetnames}")

        ws = wb[sheet_name]
        rows_out = []
        parse = ExcelReader._parse_xyz

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            # Skip fully empty rows
            if not any(v is not None for v in row):
                continue

            # col A = main p1,  col B = main p2
            main1 = parse(row[0] if len(row) > 0 else None)
            main2 = parse(row[1] if len(row) > 1 else None)

            if main1 is None or main2 is None:
                continue   # main coords mandatory

            lr = LayoutRow(row_num=i,
                           mx1=main1[0], my1=main1[1], mz1=main1[2],
                           mx2=main2[0], my2=main2[1], mz2=main2[2])

            # col D = horiz p1,  col E = horiz p2
            h1 = parse(row[3] if len(row) > 3 else None)
            h2 = parse(row[4] if len(row) > 4 else None)
            if h1 and h2:
                lr.hx1, lr.hy1, lr.hz1 = h1
                lr.hx2, lr.hy2, lr.hz2 = h2
                lr.has_horizontal = True

            # col G = vert p1,  col H = vert p2
            v1 = parse(row[6] if len(row) > 6 else None)
            v2 = parse(row[7] if len(row) > 7 else None)
            if v1 and v2:
                lr.vx1, lr.vy1, lr.vz1 = v1
                lr.vx2, lr.vy2, lr.vz2 = v2
                lr.has_vertical = True

            rows_out.append(lr)

        wb.close()
        return rows_out


# ──────────────────────────────────────────────
# AutoCAD COM Interface
# ──────────────────────────────────────────────

class AutoCADInterface:

    @staticmethod
    def inject(lisp_path: str) -> tuple[bool, str]:
        try:
            import win32com.client
            acad = win32com.client.GetActiveObject("AutoCAD.Application")
            doc = acad.ActiveDocument
            escaped = lisp_path.replace("\\", "\\\\")
            doc.SendCommand(f'(load "{escaped}") \n')
            return True, "LISP loaded into AutoCAD successfully."
        except ImportError:
            return False, "pywin32 not installed.\nRun: pip install pywin32"
        except Exception as e:
            return False, f"AutoCAD COM error: {e}"


# ──────────────────────────────────────────────
# Main GUI
# ──────────────────────────────────────────────

DARK_BG     = "#1E2128"
PANEL_BG    = "#252830"
ACCENT      = "#00B4D8"
ACCENT2     = "#FF6B35"
TEXT_MAIN   = "#E8EAF0"
TEXT_DIM    = "#7A8090"
BORDER      = "#333844"
SUCCESS     = "#4CAF50"
WARNING     = "#FFC107"
FONT_MONO   = ("Consolas", 9)
FONT_UI     = ("Segoe UI", 9)
FONT_LABEL  = ("Segoe UI", 8)
FONT_TITLE  = ("Segoe UI Semibold", 10)


class PantherApp:

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Panther  ·  AutoCAD Layout Generator")
        self.root.geometry("1000x720")
        self.root.configure(bg=DARK_BG)
        self.root.resizable(True, True)

        self.rows: list[LayoutRow] = []
        self.lisp_gen = LispGenerator()
        self.acad = AutoCADInterface()
        self.excel_path: Optional[str] = None

        self._style()
        self._build_ui()

    # ── Style ──────────────────────────────────

    def _style(self):
        s = ttk.Style()
        s.theme_use("clam")

        s.configure(".", background=DARK_BG, foreground=TEXT_MAIN,
                    font=FONT_UI, borderwidth=0, focuscolor=ACCENT)

        s.configure("TFrame", background=DARK_BG)
        s.configure("Panel.TFrame", background=PANEL_BG)

        s.configure("TLabel", background=DARK_BG, foreground=TEXT_MAIN, font=FONT_UI)
        s.configure("Dim.TLabel", background=PANEL_BG, foreground=TEXT_DIM, font=FONT_LABEL)
        s.configure("Title.TLabel", background=PANEL_BG, foreground=TEXT_MAIN, font=FONT_TITLE)
        s.configure("Accent.TLabel", background=PANEL_BG, foreground=ACCENT, font=FONT_TITLE)

        s.configure("TLabelframe", background=PANEL_BG, foreground=TEXT_DIM,
                    font=FONT_LABEL, bordercolor=BORDER, relief="flat")
        s.configure("TLabelframe.Label", background=PANEL_BG, foreground=TEXT_DIM, font=FONT_LABEL)

        s.configure("TEntry", fieldbackground="#2A2E38", foreground=TEXT_MAIN,
                    insertcolor=ACCENT, bordercolor=BORDER, relief="flat", padding=4)
        s.map("TEntry", fieldbackground=[("focus", "#2E3340")])

        s.configure("TCombobox", fieldbackground="#2A2E38", foreground=TEXT_MAIN,
                    background="#2A2E38", arrowcolor=ACCENT, bordercolor=BORDER)
        s.map("TCombobox", fieldbackground=[("readonly", "#2A2E38")])

        s.configure("Primary.TButton", background=ACCENT, foreground="#0A0E14",
                    font=("Segoe UI Semibold", 9), padding=(12, 6), relief="flat", borderwidth=0)
        s.map("Primary.TButton",
              background=[("active", "#0094B8"), ("pressed", "#007A9A")])

        s.configure("Action.TButton", background=ACCENT2, foreground="#0A0E14",
                    font=("Segoe UI Semibold", 9), padding=(12, 6), relief="flat", borderwidth=0)
        s.map("Action.TButton",
              background=[("active", "#E05520"), ("pressed", "#C04010")])

        s.configure("Ghost.TButton", background=PANEL_BG, foreground=TEXT_DIM,
                    font=FONT_UI, padding=(10, 5), relief="flat", borderwidth=1,
                    bordercolor=BORDER)
        s.map("Ghost.TButton",
              foreground=[("active", TEXT_MAIN)],
              background=[("active", "#2E3340")])

        s.configure("Treeview", background="#1A1E26", foreground=TEXT_MAIN,
                    fieldbackground="#1A1E26", bordercolor=BORDER, rowheight=24, font=FONT_UI)
        s.configure("Treeview.Heading", background="#252830", foreground=TEXT_DIM,
                    font=FONT_LABEL, relief="flat", bordercolor=BORDER)
        s.map("Treeview",
              background=[("selected", "#1A3A4A")],
              foreground=[("selected", ACCENT)])

        s.configure("Horizontal.TProgressbar", background=ACCENT,
                    troughcolor="#2A2E38", bordercolor=BORDER)

    # ── UI Build ───────────────────────────────

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg="#161920", height=48)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="⚡  PANTHER", bg="#161920", fg=ACCENT,
                 font=("Segoe UI Semibold", 13), padx=16).pack(side="left", pady=12)
        tk.Label(hdr, text="AutoCAD Layout Generator  v2", bg="#161920",
                 fg=TEXT_DIM, font=FONT_LABEL).pack(side="left", pady=12)

        # Main paned layout
        pw = tk.PanedWindow(self.root, orient="horizontal", bg=DARK_BG,
                            sashwidth=4, sashrelief="flat", sashpad=2)
        pw.pack(fill="both", expand=True, padx=8, pady=8)

        # Left panel
        left = ttk.Frame(pw, style="Panel.TFrame")
        pw.add(left, minsize=340)

        # Right panel (preview)
        right = ttk.Frame(pw, style="Panel.TFrame")
        pw.add(right, minsize=300)

        self._build_left(left)
        self._build_right(right)

    def _build_left(self, parent):
        parent.columnconfigure(0, weight=1)

        # ── Source Section ──
        src = ttk.LabelFrame(parent, text="  DATA SOURCE", padding=12)
        src.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        src.columnconfigure(1, weight=1)

        self.excel_label = ttk.Label(src, text="No file loaded", style="Dim.TLabel")
        self.excel_label.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        ttk.Button(src, text="Load Excel (PASTE sheet)",
                   style="Primary.TButton",
                   command=self._load_excel).grid(row=1, column=0, sticky="w")

        self.row_count_var = tk.StringVar(value="")
        ttk.Label(src, textvariable=self.row_count_var, style="Dim.TLabel").grid(
            row=1, column=1, sticky="e", padx=8)

        # ── Scale Section ──
        scl = ttk.LabelFrame(parent, text="  SCALE", padding=12)
        scl.grid(row=1, column=0, sticky="ew", padx=8, pady=4)
        scl.columnconfigure(1, weight=1)

        self.scale_var = tk.StringVar(value="1:100")
        for i, opt in enumerate(["1:50", "1:100", "Custom"]):
            rb = ttk.Radiobutton(scl, text=opt, value=opt,
                                 variable=self.scale_var,
                                 command=self._on_scale_change)
            rb.grid(row=0, column=i, padx=(0, 12), sticky="w")

        ttk.Label(scl, text="1 :", style="Dim.TLabel").grid(
            row=1, column=0, sticky="e", pady=(8, 0))
        self.custom_scale_var = tk.StringVar(value="100")
        self.custom_entry = ttk.Entry(scl, textvariable=self.custom_scale_var,
                                      width=10, state="disabled")
        self.custom_entry.grid(row=1, column=1, sticky="w", padx=(4, 0), pady=(8, 0))
        ttk.Label(scl, text="(paper units per model unit)",
                  style="Dim.TLabel").grid(row=1, column=2, sticky="w", padx=6, pady=(8, 0))

        # ── Layout Prefix ──
        pfx = ttk.LabelFrame(parent, text="  LAYOUT PREFIX", padding=12)
        pfx.grid(row=2, column=0, sticky="ew", padx=8, pady=4)
        pfx.columnconfigure(1, weight=1)

        self.prefix_var = tk.StringVar(value="Layout")
        ttk.Entry(pfx, textvariable=self.prefix_var, width=20).grid(
            row=0, column=0, sticky="w")
        ttk.Label(pfx, text="e.g.  Layout_1, Layout_2 …",
                  style="Dim.TLabel").grid(row=0, column=1, sticky="w", padx=10)

        # ── Coordinate Table ──
        tbl = ttk.LabelFrame(parent, text="  COORDINATE ROWS", padding=8)
        tbl.grid(row=3, column=0, sticky="nsew", padx=8, pady=4)
        tbl.columnconfigure(0, weight=1)
        tbl.rowconfigure(0, weight=1)
        parent.rowconfigure(3, weight=1)

        cols = ("#", "Main (X1,Y1)→(X2,Y2)", "Orient", "Callout")
        self.tree = ttk.Treeview(tbl, columns=cols, show="headings", height=10)
        widths = [30, 220, 60, 70]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=w)

        vsb = ttk.Scrollbar(tbl, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        # ── Action Buttons ──
        act = ttk.Frame(parent, style="Panel.TFrame")
        act.grid(row=4, column=0, sticky="ew", padx=8, pady=(4, 8))

        ttk.Button(act, text="Generate .LSP",
                   style="Primary.TButton",
                   command=self._generate_lsp).pack(side="left", padx=(0, 6))

        ttk.Button(act, text="⚡ Run in AutoCAD",
                   style="Action.TButton",
                   command=self._run_in_autocad).pack(side="left", padx=(0, 6))

        ttk.Button(act, text="Refresh Preview",
                   style="Ghost.TButton",
                   command=self._refresh_preview).pack(side="left")

        # ── Status bar ──
        self.status_var = tk.StringVar(value="Ready")
        self.status_bar = tk.Label(parent, textvariable=self.status_var,
                                   bg="#161920", fg=TEXT_DIM,
                                   font=FONT_LABEL, anchor="w", padx=10, pady=4)
        self.status_bar.grid(row=5, column=0, sticky="ew")

    def _build_right(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        hdr = ttk.Label(parent, text="LISP PREVIEW", style="Dim.TLabel")
        hdr.grid(row=0, column=0, sticky="w", padx=10, pady=(10, 4))

        self.preview = tk.Text(parent, bg="#0E1117", fg="#A8C8E8",
                               font=FONT_MONO, wrap="none",
                               insertbackground=ACCENT,
                               selectbackground="#1A3A4A",
                               relief="flat", borderwidth=0,
                               padx=12, pady=10)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.preview.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self.preview.xview)
        self.preview.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.preview.grid(row=1, column=0, sticky="nsew", padx=(8, 0), pady=(0, 0))
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew", padx=(8, 0))

        # Syntax coloring tags
        self.preview.tag_configure("comment", foreground="#5C7A5C")
        self.preview.tag_configure("keyword", foreground=ACCENT)
        self.preview.tag_configure("string",  foreground="#CE9178")
        self.preview.tag_configure("number",  foreground="#B5CEA8")

        copy_btn = ttk.Button(parent, text="Copy to clipboard",
                              style="Ghost.TButton",
                              command=self._copy_preview)
        copy_btn.grid(row=3, column=0, sticky="e", padx=8, pady=6)

    # ── Helpers ────────────────────────────────

    def _get_scale(self) -> Optional[float]:
        sel = self.scale_var.get()
        if sel == "1:50":
            return 50.0
        elif sel == "1:100":
            return 100.0
        else:
            try:
                v = float(self.custom_scale_var.get())
                if v <= 0:
                    raise ValueError
                return v
            except ValueError:
                messagebox.showerror("Invalid Scale",
                    "Enter a positive number for custom scale.")
                return None

    def _on_scale_change(self):
        if self.scale_var.get() == "Custom":
            self.custom_entry.configure(state="normal")
        else:
            self.custom_entry.configure(state="disabled")
        self._refresh_preview()

    def _set_status(self, msg: str, color: str = TEXT_DIM):
        self.status_var.set(msg)
        self.status_bar.configure(fg=color)
        self.root.update_idletasks()

    def _refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        for lr in self.rows:
            orient = "HORIZ" if lr.is_horizontal else "VERT"
            has_c = "✓" if lr.callout_coords else "—"
            coords = f"({lr.mx1:.1f},{lr.my1:.1f})→({lr.mx2:.1f},{lr.my2:.1f})"
            self.tree.insert("", "end", values=(lr.row_num, coords, orient, has_c))

    def _refresh_preview(self):
        scale = self._get_scale()
        if scale is None or not self.rows:
            self.preview.delete("1.0", "end")
            self.preview.insert("1.0", "; No data loaded yet.\n; Load an Excel file to generate LISP.")
            return
        lisp = self.lisp_gen.generate(self.rows, scale, self.prefix_var.get())
        self.preview.delete("1.0", "end")
        self.preview.insert("1.0", lisp)
        self._highlight_preview()

    def _highlight_preview(self):
        content = self.preview.get("1.0", "end")
        self.preview.tag_remove("comment", "1.0", "end")
        self.preview.tag_remove("keyword", "1.0", "end")
        self.preview.tag_remove("string",  "1.0", "end")
        self.preview.tag_remove("number",  "1.0", "end")

        import re
        for match in re.finditer(r";[^\n]*", content):
            s = f"1.0+{match.start()}c"
            e = f"1.0+{match.end()}c"
            self.preview.tag_add("comment", s, e)
        for match in re.finditer(r'"[^"]*"', content):
            s = f"1.0+{match.start()}c"
            e = f"1.0+{match.end()}c"
            self.preview.tag_add("string", s, e)
        for kw in ["defun", "command", "setvar", "setq", "princ", "strcat", "itoa", "vla-get", "vlax-get"]:
            for match in re.finditer(rf'\b{kw}\b', content):
                s = f"1.0+{match.start()}c"
                e = f"1.0+{match.end()}c"
                self.preview.tag_add("keyword", s, e)
        for match in re.finditer(r'\b\d+\.\d+\b', content):
            s = f"1.0+{match.start()}c"
            e = f"1.0+{match.end()}c"
            self.preview.tag_add("number", s, e)

    # ── Actions ────────────────────────────────

    def _load_excel(self):
        fp = filedialog.askopenfilename(
            title="Select input file (Excel or CSV)",
            filetypes=[("Excel / CSV", "*.xlsx *.xlsm *.csv"), ("All Files", "*.*")]
        )
        if not fp:
            return
        try:
            if fp.lower().endswith(".csv"):
                self.rows = ExcelReader.read_csv(fp)
            else:
                self.rows = ExcelReader.read_paste_sheet(fp)
            self.excel_path = fp
            fname = Path(fp).name
            self.excel_label.configure(text=fname, foreground=ACCENT)
            self.row_count_var.set(f"{len(self.rows)} row(s)")
            self._refresh_tree()
            self._refresh_preview()
            self._set_status(f"Loaded {len(self.rows)} layout row(s) from {fname}", SUCCESS)
        except Exception as e:
            messagebox.showerror("Load Error", str(e))
            self._set_status(f"Error: {e}", ACCENT2)

    def _generate_lsp(self) -> Optional[str]:
        if not self.rows:
            messagebox.showwarning("No Data", "Load an Excel file first.")
            return None
        scale = self._get_scale()
        if scale is None:
            return None

        fp = filedialog.asksaveasfilename(
            title="Save LISP file",
            defaultextension=".lsp",
            initialfile="panther_layouts.lsp",
            filetypes=[("AutoLISP Files", "*.lsp"), ("All Files", "*.*")]
        )
        if not fp:
            return None

        lisp = self.lisp_gen.generate(self.rows, scale, self.prefix_var.get())
        with open(fp, "w") as f:
            f.write(lisp)
        self._set_status(f"Saved: {fp}", SUCCESS)
        messagebox.showinfo("Saved", f"LISP file saved to:\n{fp}")
        return fp

    def _run_in_autocad(self):
        if not self.rows:
            messagebox.showwarning("No Data", "Load an Excel file first.")
            return
        scale = self._get_scale()
        if scale is None:
            return

        lisp = self.lisp_gen.generate(self.rows, scale, self.prefix_var.get())

        with tempfile.NamedTemporaryFile(mode="w", suffix=".lsp",
                                         delete=False, prefix="panther_") as f:
            f.write(lisp)
            tmp = f.name

        self._set_status("Connecting to AutoCAD…", WARNING)
        ok, msg = self.acad.inject(tmp)

        if ok:
            self._set_status(msg, SUCCESS)
            messagebox.showinfo("AutoCAD", msg)
        else:
            self._set_status(f"COM failed: {msg}", ACCENT2)
            if messagebox.askyesno("AutoCAD Not Reachable",
                    f"{msg}\n\nWould you like to save the LISP file manually instead?"):
                self._generate_lsp()

    def _copy_preview(self):
        content = self.preview.get("1.0", "end")
        self.root.clipboard_clear()
        self.root.clipboard_append(content)
        self._set_status("LISP copied to clipboard.", SUCCESS)


# ──────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────

def main():
    root = tk.Tk()
    app = PantherApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
